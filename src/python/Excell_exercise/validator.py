import openpyxl
import datetime

def validator(file):
    workbook = openpyxl.open(file)
    if not file.endswith(".xlsx"):
        return False
    if "Customer" not in workbook.sheetnames or "Invoices" not in workbook.sheetnames or "Invoice Line Items" not in workbook.sheetnames or "Product Table" not in workbook.sheetnames:
        return False

    for row in workbook["Customer"]:
        if row[0].value == None:
            break
        if row[0].value == "ID":
            continue
        if not int(row[0].value) is not int:
            return False
    for row in workbook["Invoices"]:
        if row[0].value == None:
            break
        if row[0].value == "ID":
            continue
        if type(row[0].value) is not int or type(row[1].value) is not datetime.datetime or type(row[2].value) is not int:
            return False
    for row in workbook["Invoice Line Items"]:
        if row[0].value == None:
            break
        if row[0].value == "ID":
            continue
        if type(row[0].value) is not int or type(row[1].value) is not int or type(row[2].value) is not int or type(row[3].value) is not int:
            return False
    for row in workbook["Product Table"]:
        if row[0].value == None:
            break
        if row[0].value == "ID":
            continue
        if type(row[0].value) is not int or type(row[1].value) is not str or type(row[2].value) is not int and type(row[2].value) is not float:
            return False        
    return True